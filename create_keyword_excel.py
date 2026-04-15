import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

path = r'C:\Users\AudreyBegin\.claude\projects\c--Programmation\2f04bae4-1061-4120-a49c-ea9bfe68c66f\tool-results\mcp-dataforseo-dataforseo_research_keywords-1776285743276.txt'
with open(path, encoding='utf-8') as f:
    raw = json.load(f)

keywords = raw['keywords']

wb = Workbook()
ws = wb.active
ws.title = "Mots-clés Sommeil"

headers = ["#", "Mot-clé", "Volume mensuel", "Compétition", "Index compétition", "CPC ($)", "Difficulté", "Intention", "Sources"]

thin = Side(style="thin", color="B8CCE4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1F4E79")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

row_font = Font(name="Arial", size=10)
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

fill_even = PatternFill("solid", start_color="DCE6F1")
fill_odd = PatternFill("solid", start_color="FFFFFF")

fill_low    = PatternFill("solid", start_color="C6EFCE")
fill_medium = PatternFill("solid", start_color="FFEB9C")
fill_high   = PatternFill("solid", start_color="FFC7CE")
font_low    = Font(name="Arial", size=10, color="276221", bold=True)
font_medium = Font(name="Arial", size=10, color="9C6500", bold=True)
font_high   = Font(name="Arial", size=10, color="9C0006", bold=True)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
ws.row_dimensions[1].height = 30

for row_idx, kw in enumerate(keywords, 2):
    fill_row = fill_even if row_idx % 2 == 0 else fill_odd

    sources = kw.get('sources', [])
    sources_str = ", ".join(sources) if sources else ""

    values = [
        row_idx - 1,
        kw.get('keyword', ''),
        kw.get('search_volume') if kw.get('search_volume') is not None else '-',
        kw.get('competition_level') or '-',
        kw.get('competition_index') if kw.get('competition_index') is not None else '-',
        kw.get('cpc') if kw.get('cpc') is not None else '-',
        kw.get('keyword_difficulty') if kw.get('keyword_difficulty') is not None else '-',
        kw.get('main_intent') or '-',
        sources_str,
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        cell.fill = fill_row

        if col_idx == 2:
            cell.font = row_font
            cell.alignment = align_left
        elif col_idx == 4:
            comp = kw.get('competition_level')
            if comp == "LOW":
                cell.font = font_low
                cell.fill = fill_low
            elif comp == "MEDIUM":
                cell.font = font_medium
                cell.fill = fill_medium
            elif comp == "HIGH":
                cell.font = font_high
                cell.fill = fill_high
            else:
                cell.font = row_font
            cell.alignment = align_center
        elif col_idx == 6 and isinstance(val, float):
            cell.font = row_font
            cell.alignment = align_center
            cell.number_format = '#,##0.00'
        else:
            cell.font = row_font
            cell.alignment = align_center

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 12
ws.column_dimensions["H"].width = 16
ws.column_dimensions["I"].width = 28

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{len(keywords) + 1}"

# Summary sheet
ws2 = wb.create_sheet("Résumé")
summary = [
    ("Statistiques", "Valeur"),
    ("Mots-clés bruts extraits", raw.get('total_raw', '')),
    ("Après déduplication", raw.get('total_after_dedup', '')),
    ("Après filtre marques", raw.get('total_after_brand_filter', '')),
    ("Dans ce tableau", len(keywords)),
    ("Source : Google Ads", raw.get('sources_summary', {}).get('google_ads', '')),
    ("Source : Ideas", raw.get('sources_summary', {}).get('ideas', '')),
    ("Source : Suggestions", raw.get('sources_summary', {}).get('suggestions', '')),
    ("Marché cible", "Canada — Français (fr-CA)"),
    ("Date d'extraction", "2026-04-15"),
]

for row_idx, (label, value) in enumerate(summary, 1):
    c1 = ws2.cell(row=row_idx, column=1, value=label)
    c2 = ws2.cell(row=row_idx, column=2, value=value)
    for c in [c1, c2]:
        c.border = border
        c.alignment = align_left
        if row_idx == 1:
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = header_fill
        else:
            c.font = row_font
            c.fill = fill_even if row_idx % 2 == 0 else fill_odd

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 35

out = r'C:\Programmation\recherche_mots_cles_sommeil.xlsx'
wb.save(out)
print(f"Fichier sauvegardé : {out} ({len(keywords)} mots-clés)")
