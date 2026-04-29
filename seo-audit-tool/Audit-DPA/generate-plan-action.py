from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Couleurs ────────────────────────────────────────────────────────────────
NAVY       = "1A3A5C"
NAVY_LIGHT = "E8F0F7"
WHITE      = "FFFFFF"
GREY_ROW   = "F5F7FA"
GREY_BORDER= "C5CDD8"

COL_CRIT   = ("C0392B", "FDECEA")   # rouge  — texte, fond
COL_ELEV   = ("B8560A", "FDF0EA")   # orange — texte, fond
COL_MOY    = ("8B6914", "FEF6E7")   # jaune  — texte, fond
COL_FAIB   = ("1E7B5A", "E6F4EF")   # vert   — texte, fond

PRIO_MAP = {
    "Critique": COL_CRIT,
    "Élevée":   COL_ELEV,
    "Moyenne":  COL_MOY,
    "Faible":   COL_FAIB,
}

# ── Helpers ─────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def border():
    s = Side(style="thin", color=GREY_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, title, subtitle):
    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
    ws["A1"].fill = fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="7FAACC")
    ws["A2"].fill = fill(NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

def write_col_headers(ws, row):
    headers = ["#", "Action principale", "Description", "Impact attendu", "Priorité"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = fill(NAVY_LIGHT.replace("E8F0F7", "2E6BA8"))
        c.alignment = Alignment(horizontal="center" if col == 1 else "left",
                                vertical="center", wrap_text=True, indent=0 if col == 1 else 1)
        c.border = border()
    ws.row_dimensions[row].height = 22

def write_row(ws, row, num, action, description, impact, priorite):
    text_color, bg_color = PRIO_MAP.get(priorite, ("000000", WHITE))
    row_bg = GREY_ROW if row % 2 == 0 else WHITE

    data = [num, action, description, impact, priorite]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border()
        c.alignment = Alignment(horizontal="center" if col == 1 else "left",
                                vertical="top", wrap_text=True, indent=0 if col == 1 else 1)
        if col == 5:
            c.font = Font(name="Arial", size=10, bold=True, color=text_color)
            c.fill = fill(bg_color)
        else:
            c.font = Font(name="Arial", size=10, bold=(col == 2), color="1E293B")
            c.fill = fill(row_bg)
    ws.row_dimensions[row].height = 70

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ── Données Phase 1 ─────────────────────────────────────────────────────────
phase1 = [
    (1,  "Corriger les redirections brisées",
         "Obtenir la table de redirections complète de l'agence ayant géré la migration. Identifier les URLs sources et leurs équivalents actuels. Corriger les redirections mal configurées ou absentes. Revalider via Google Search Console.",
         "Récupération partielle et structurelle du trafic perdu depuis janvier 2026. C'est la cause #1 de la chute de -46,6%.",
         "Critique"),

    (2,  "Réactiver l'article « 7 sons FR »",
         "Changer la balise noindex en index, follow sur la version française de l'article blogue. Soumettre l'URL pour réindexation via Google Search Console > Inspection d'URL.",
         "Récupération du trafic francophone sous 2 à 4 semaines. CTR version FR = 7,6% vs 0,8% EN — les francophones cliquaient 10× plus.",
         "Critique"),

    (3,  "Régénérer le sitemap XML",
         "Retirer du sitemap toutes les URLs en 404, en noindex, les pages /node/, les pages /en/ et les pages de remerciement. Règle : le sitemap ne doit contenir que des pages indexables (statut 200, sans noindex).",
         "Amélioration du crawl budget. Google indexe les bonnes pages et cesse de crawler les pages mortes.",
         "Critique"),

    (4,  "Rediriger les 24+ pages /en/ en 301",
         "Rediriger chaque URL /en/ vers son équivalent /fr/. 24 pages identifiées dans l'audit (blog, réclamations, bureaux, nous joindre). Retirer toutes les URLs /en/ du sitemap.",
         "Consolidation de l'autorité accumulée sur les pages anglaises vers les pages françaises.",
         "Critique"),

    (5,  "Corriger la canonical contradictoire sur /carrieres",
         "La page de destination d'une redirection 301 pointe son canonical vers l'ancienne URL — boucle logique. Corriger pour une balise self-referencing. Corriger la typo « courtierere » dans le slug.",
         "Récupération des 685 impressions perdues à zéro depuis la migration.",
         "Critique"),

    (6,  "Créer la fiche GMB — Roxton Pond",
         "Créer et compléter la fiche Google Business Profile pour le bureau de Roxton Pond (acquisition récente sans présence locale). Adresse, téléphone, horaires, catégorie, photos.",
         "Visibilité locale immédiate pour une succursale actuellement invisible sur Google Maps.",
         "Critique"),

    (7,  "Retirer les pages /node/ et techniques du sitemap",
         "Ajouter noindex sur les 10 pages /node/ et la page /carrieres/merci. Les retirer du sitemap. Si certaines redirigent encore, pointer vers une page pertinente.",
         "Qualité du crawl améliorée. Google concentre ses ressources sur les pages utiles.",
         "Élevée"),

    (8,  "Corriger les Core Web Vitals (CLS 0,29 — seuil 0,1)",
         "Ajouter width et height sur toutes les images. Réserver l'espace du carrousel hero avec min-height. Supprimer ~755 Ko de JS inutilisé et ~195 Ko de CSS inutilisé. Précharger l'image LCP.",
         "Facteur de classement Google récupéré sur desktop. Expérience utilisateur améliorée.",
         "Élevée"),

    (9,  "Résoudre le contenu dupliqué (dommage par eau)",
         "Rediriger /fr/dommage-par-eau-en-6-etapes en 301 vers l'article de blogue : /fr/a-propos/blogue/allo-courtier/quoi-faire-en-cas-de-dommages-par-leau-en-6-etapes.",
         "L'autorité de la page landing se consolide dans l'article blogue. Google sait quelle version valoriser.",
         "Élevée"),

    (10, "Optimiser la fiche GMB — Brossard (8e position)",
         "Ajouter des photos récentes, rédiger une description avec mots-clés locaux, publier des posts réguliers, répondre à tous les avis. Viser la complétude à 100% de la fiche.",
         "Amélioration de la position actuelle (8e) vers le top 3 local.",
         "Élevée"),

    (11, "Rédiger les 38 méta-descriptions manquantes",
         "Réécrire les méta-descriptions de toutes les pages de services. Format : mot-clé cible + bénéfice client + appel à l'action. 120-160 caractères. Modèle fourni dans l'audit.",
         "Amélioration du taux de clic dans les SERP. Différenciation dans les résultats Google.",
         "Moyenne"),

    (12, "Résoudre les 22 pages carrières en 403",
         "Rediriger les URLs de carrières retournant une erreur 403 (accès interdit) en 301 vers la page carrières principale ou vers le poste actif équivalent.",
         "Fin de l'exposition de pages d'erreur à Google. Amélioration de l'expérience des visiteurs.",
         "Moyenne"),

    (13, "Corriger les ~25 images et PDFs brisés post-migration",
         "Retrouver les ressources 404 dans les backups pre-migration. Re-uploader aux mêmes chemins. Si introuvables, retirer les références dans les articles. Ne pas créer de redirections 301 sur les images.",
         "Expérience utilisateur améliorée dans le blogue. Fin des ressources cassées visibles.",
         "Moyenne"),

    (14, "Lancer une campagne d'avis Google — Drummondville",
         "Programme de collecte d'avis auprès des clients actuels du bureau de Drummondville. Objectif : passer de 53 à 150 avis en 6 mois.",
         "Renforcement de la crédibilité locale. Amélioration du classement GMB à Drummondville.",
         "Moyenne"),

    (15, "Désavouer les liens de scraping B2B",
         "Créer un fichier de désaveu via Google Search Console pour les domaines identifiés (clodura, success.ai, salezshark). Soumettre via l'outil de désaveu Google.",
         "Assainissement du profil de backlinks. Impact marginal mais bonne hygiène SEO.",
         "Faible"),
]

# ── Données Phase 2 ─────────────────────────────────────────────────────────
phase2 = [
    (1,  "Réécrire les H1 de toutes les pages de services",
         "Remplacer les H1 navigationnels (« Automobile », « Ma maison ») par des H1 orientés mots-clés. Règle unique : le H1 doit correspondre à ce que quelqu'un taperait dans Google. Ex : « Assurance automobile au Québec | Courtier DPA ».",
         "Condition nécessaire au positionnement organique. Sans H1 correct, aucune autre optimisation ne peut pleinement fonctionner.",
         "Critique"),

    (2,  "Enrichir les pages de services en profondeur",
         "Passer de 150-400 mots à 1 500-2 000 mots structurés par page. Chaque page répond aux questions que l'utilisateur se pose avant de demander une soumission. Architecture H2/H3 détaillée dans l'audit pour auto, maison, agricole et biens/RC.",
         "Positionnement sur requêtes à volume : auto (2 600/mois), habitation (2 100/mois). Pages actuellement invisibles sur la longue traîne.",
         "Critique"),

    (3,  "Ajouter des sections FAQ avec schéma FAQPage",
         "Intégrer une section FAQ structurée sur chaque page de service (minimum 5 Q/R). Ajouter le schéma FAQPage (données structurées JSON-LD) sur chaque section FAQ.",
         "Éligibilité aux featured snippets Google. Réponse directe aux intentions de recherche de type question.",
         "Critique"),

    (4,  "Construire un maillage interne structuré",
         "Chaque page produit pointe vers les pages connexes dans le corps du texte. Règles : auto ↔ maison (combo Synchro), auto → assistance routière → pardon des accidents, maison → dommage par eau. Chaque article de blogue pointe vers la page de service correspondante.",
         "Transfert d'autorité entre pages. Amélioration du classement des pages commerciales. Réduction du taux de rebond.",
         "Critique"),

    (5,  "Mettre en place la stratégie de contenu blogue (clusters)",
         "Organiser le blogue en 4 clusters : auto, habitation, entreprise, courtier. Publier minimum 2 articles/mois. Chaque article répond à une question cherchée sur Google et pointe vers une page de service avec CTA. Détail des clusters dans l'audit.",
         "Génération progressive de trafic hors marque. Cible : passer de 1% à 15% de trafic hors marque en 12 mois.",
         "Critique"),

    (6,  "Créer les pages produits manquantes",
         "Pages prioritaires à créer : /assurances-entreprises/produits-specifiques/erreurs-et-omissions, /assurances-entreprises/pme (page hub), /travailleurs-autonomes (page hub), /vehicule-electrique. Présents chez 3-4 concurrents sur 4.",
         "Couverture des gaps produits vs concurrents. Captation de trafic sur des requêtes sans concurrence actuelle chez DPA.",
         "Élevée"),

    (7,  "Intégrer les données structurées dès le développement",
         "Configurer LocalBusiness sur chaque page de bureau (adresse, téléphone, GPS, horaires). Configurer Organization complet sur l'accueil. Configurer FAQPage sur toutes les pages produits avec section FAQ.",
         "Éligibilité aux rich results Google. Amélioration du SEO local. Renforcement de l'autorité de domaine.",
         "Élevée"),

    (8,  "Recréer la page /notre-equipe",
         "Page avec photos terrain (pas portraits de bureau), noms, titres et expertises des courtiers. Mario Poirier en avant pour l'agricole. Citations, spécialisations régionales, années d'expérience.",
         "Signal E-E-A-T fort pour un secteur classé YMYL (Your Money or Your Life) par Google. Facteur de confiance pour les nouveaux visiteurs.",
         "Élevée"),

    (9,  "Implémenter les pages locales par ville",
         "Créer une page dédiée par ville principale : « Assurance auto à Saint-Hyacinthe », « Courtier assurance Drummondville », « Assurance habitation Brossard ». DPA est 1er GMB à St-Hyacinthe et Drummondville — aucune page locale n'existe.",
         "Captation du trafic local longue traîne. Renforcement du SEO local en complément des fiches GMB.",
         "Élevée"),

    (10, "Ajouter un H1 à la page d'accueil",
         "La page d'accueil n'a pas de H1 (bug template). Rédiger un H1 avec le mot-clé principal de l'accueil : ex. « Courtier en assurance au Québec — Groupe DPA ».",
         "Signal de pertinence rétabli sur la page la plus visitée du site.",
         "Moyenne"),

    (11, "Stratégie d'acquisition de liens (off-site)",
         "Identifier 20-30 cibles de liens : associations professionnelles régionales (UPA, chambres de commerce), médias économiques Montérégie/Centre-du-Québec, partenaires agricoles. Objectif : DA de 21 à 30+ en 12 mois.",
         "Amélioration progressive de l'autorité de domaine. Impact sur le positionnement de toutes les pages du site.",
         "Moyenne"),

    (12, "Configurer les balises title selon le modèle recommandé",
         "Format cible pour toutes les pages produits : « Assurance [produit] au Québec | DPA Assurances ». Le mot-clé cible en premier, la marque en suffixe. Intégrer dès le gabarit du nouveau site.",
         "Amélioration du taux de clic et du positionnement sur les mots-clés produits.",
         "Moyenne"),

    (13, "Intégrer les pages orphelines dans la navigation",
         "Lors de la refonte, mapper les 17+ pages actuellement orphelines (sans lien interne). Les intégrer dans la navigation ou le maillage. Supprimer ou rediriger celles sans valeur.",
         "Consolidation de l'autorité du site. Fin des pages invisibles à Google malgré leur présence dans le sitemap.",
         "Moyenne"),
]

# ── Construction des feuilles ────────────────────────────────────────────────
def build_sheet(ws, title, subtitle, data):
    write_header(ws, title, subtitle)
    write_col_headers(ws, 3)
    for i, row_data in enumerate(data):
        write_row(ws, i + 4, *row_data)
    total_rows = len(data) + 4

    # Ligne de légende
    ws.row_dimensions[total_rows].height = 16
    leg_row = total_rows + 1
    ws.merge_cells(f"A{leg_row}:E{leg_row}")
    ws[f"A{leg_row}"] = (
        "Légende — Priorité :   🔴 Critique = action immédiate, impact direct sur le trafic   "
        "🟠 Élevée = important, exécuter dans le mois   "
        "🟡 Moyenne = à planifier   "
        "🟢 Faible = hygiène SEO"
    )
    ws[f"A{leg_row}"].font = Font(name="Arial", size=9, italic=True, color="64748B")
    ws[f"A{leg_row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[leg_row].height = 18

    set_col_widths(ws, [5, 30, 52, 38, 13])
    ws.freeze_panes = "A4"

# ── Feuille 1 — Phase 1 ──────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Phase 1 — Site actuel"
build_sheet(
    ws1,
    "Plan d'action — Phase 1 : Site actuel",
    "Corrections techniques à impact immédiat — indépendantes de la refonte",
    phase1
)

# ── Feuille 2 — Phase 2 ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("Phase 2 — Nouveau site")
build_sheet(
    ws2,
    "Plan d'action — Phase 2 : Nouveau site",
    "Recommandations à intégrer dès la conception de la refonte",
    phase2
)

# ── Feuille 3 — Vue consolidée ───────────────────────────────────────────────
ws3 = wb.create_sheet("Résumé consolidé")
ws3.merge_cells("A1:F1")
ws3["A1"] = "Plan d'action SEO — Groupe DPA Assurances — Résumé consolidé"
ws3["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws3["A1"].fill = fill(NAVY)
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:F2")
ws3["A2"] = "Audit SEO — Avril 2026"
ws3["A2"].font = Font(name="Arial", size=10, italic=True, color="7FAACC")
ws3["A2"].fill = fill(NAVY)
ws3["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.row_dimensions[2].height = 18

headers_c = ["#", "Phase", "Action principale", "Description", "Impact attendu", "Priorité"]
for col, h in enumerate(headers_c, 1):
    c = ws3.cell(row=3, column=col, value=h)
    c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    c.fill = fill("2E6BA8")
    c.alignment = Alignment(horizontal="center" if col <= 2 else "left",
                             vertical="center", wrap_text=True, indent=0 if col <= 2 else 1)
    c.border = border()
ws3.row_dimensions[3].height = 22

all_rows = [(r[0], "Phase 1 — Site actuel", r[1], r[2], r[3], r[4]) for r in phase1] + \
           [(r[0], "Phase 2 — Nouveau site", r[1], r[2], r[3], r[4]) for r in phase2]

for i, (num, phase, action, desc, impact, prio) in enumerate(all_rows):
    row = i + 4
    text_color, bg_color = PRIO_MAP.get(prio, ("000000", WHITE))
    row_bg = GREY_ROW if i % 2 == 0 else WHITE
    phase_bg = "EBF2FA" if "Phase 1" in phase else "E8F5F0"

    vals = [num, phase, action, desc, impact, prio]
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.border = border()
        c.alignment = Alignment(
            horizontal="center" if col <= 2 else "left",
            vertical="top", wrap_text=True,
            indent=0 if col <= 2 else 1
        )
        if col == 6:
            c.font = Font(name="Arial", size=10, bold=True, color=text_color)
            c.fill = fill(bg_color)
        elif col == 2:
            c.font = Font(name="Arial", size=9, bold=True, color="1A3A5C" if "Phase 1" in phase else "1E7B5A")
            c.fill = fill(phase_bg)
        else:
            c.font = Font(name="Arial", size=10, bold=(col == 3), color="1E293B")
            c.fill = fill(row_bg)
    ws3.row_dimensions[row].height = 65

ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 28
ws3.column_dimensions["D"].width = 48
ws3.column_dimensions["E"].width = 34
ws3.column_dimensions["F"].width = 13
ws3.freeze_panes = "A4"

# ── Sauvegarde ───────────────────────────────────────────────────────────────
out = "C:/Programmation/seo-audit-tool/Audit-DPA/Plan-Action-SEO-DPA.xlsx"
wb.save(out)
print(f"Fichier créé : {out}")
